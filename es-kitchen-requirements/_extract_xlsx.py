"""
One-shot extractor: function_list.xlsx → .claude/context/business-flows/*.md

Rules:
- Skip sheets 11, 21, 22, 23 (backup / phase-2 brSE)
- Per-domain markdown for BF_* sheets → domains/<slug>.md
- Function List, Overview, Screen Code Rule, Business Flow → top-level md
- Compact markdown tables, drop empty cols, trim long cells
"""
import openpyxl
from pathlib import Path
import re
import sys

SRC = Path('es-kitchen-requirements/function_list.xlsx')
OUT = Path('.claude/context/business-flows')

# Sheet names → output config
SKIP = {
    'backup_function_2',
    'Copy of phase 2_brse',
    'phase 2_brse',
    'backup_Function List',
}

# slug for BF_ sheets
BF_SLUG = {
    'BF_HỢP ĐỒNG Quản lý Hợp đồng': 'hop-dong',
    'BF_MENU & ORDER Quản lý Thực đơ': 'menu-order',
    'BF_GIAO HÀNG Lịch trình & Điều ': 'giao-hang-dieu-phoi',
    'BF_ĐẶT HÀNG NCC Đặt hàng Nhà cu': 'dat-hang-ncc',
    'BF_GIAO HÀNG Web Đối tác Vận ch': 'giao-hang-doi-tac',
    'BF_GIAO HÀNG App Tài xế': 'giao-hang-tai-xe',
    'BF_THANH TOÁN Thanh toán & Hoàn': 'thanh-toan',
    'BF_THU TIỀN & HÀNG HỦY': 'thu-tien-huy',
    'BF_TỒN KHO & THIẾT BỊ Quản lý T': 'ton-kho-thiet-bi',
    'BF_USER BINDING Liên kết Nhân v': 'user-binding',
    'BF_USER ENGAGEMENT Tương tác & ': 'user-engagement',
    'BF_MARKETING Giới thiệu Công ty': 'marketing',
    'BF_ĐẠI LÝ Quản lý Đại lý (Agenc': 'dai-ly',
    'BF_SYSTEM & OTHER Cấu hình Hệ t': 'system-other',
}


def cell_to_str(v):
    if v is None:
        return ''
    s = str(v).replace('\r\n', ' ').replace('\n', ' ').strip()
    # Collapse multiple spaces
    s = re.sub(r' +', ' ', s)
    # Escape pipe for markdown table
    s = s.replace('|', '\\|')
    return s


def read_rows(ws):
    """Yield non-empty rows trimmed to last non-empty col."""
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        # Strip trailing empty cells
        while row and (row[-1] is None or str(row[-1]).strip() == ''):
            row.pop()
        if not row:
            continue
        yield r, row


def find_header_row(rows, must_contain):
    """Find row index whose values include all keywords in must_contain."""
    for r, row in rows:
        vals = [str(v).strip() if v else '' for v in row]
        if all(any(kw == v for v in vals) for kw in must_contain):
            return r, row
    return None, None


def table_md(header, data_rows):
    """Return markdown table. Drop fully-empty columns."""
    # Determine col count
    ncol = max(len(header), max((len(r) for r in data_rows), default=0))
    # Find non-empty columns (header or any row has data)
    keep = []
    for c in range(ncol):
        hv = header[c] if c < len(header) else ''
        anydata = any(c < len(r) and cell_to_str(r[c]) for r in data_rows)
        if cell_to_str(hv) or anydata:
            keep.append(c)
    if not keep:
        return ''
    # Build
    def pick(row, c):
        return cell_to_str(row[c]) if c < len(row) else ''
    head_line = '| ' + ' | '.join(pick(header, c) or f'col{c+1}' for c in keep) + ' |'
    sep_line = '| ' + ' | '.join('---' for _ in keep) + ' |'
    body = []
    for row in data_rows:
        body.append('| ' + ' | '.join(pick(row, c) for c in keep) + ' |')
    return '\n'.join([head_line, sep_line] + body)


def extract_bf(ws, sheet_name, slug):
    """BF_* sheets: header at row 4 with ID/EPIC/STORY. Data from row 5."""
    rows = list(read_rows(ws))
    # Find header row: contains 'ID' and 'STORY' and 'DESCRIPTION'
    header_row_idx = None
    header_vals = None
    for r, row in rows:
        vals = [cell_to_str(v) for v in row]
        if 'ID' in vals and 'STORY' in vals and 'DESCRIPTION' in vals:
            header_row_idx = r
            header_vals = vals
            break
    if header_row_idx is None:
        # Fallback: just dump all rows
        md = [f'# {sheet_name}\n', '> Header không nhận diện được — dump raw.\n']
        for r, row in rows:
            md.append(f'- R{r}: {" | ".join(cell_to_str(v) for v in row)}')
        return '\n'.join(md)
    # Strip leading empty column from header
    while header_vals and header_vals[0] == '':
        header_vals.pop(0)
        leading = 1
    leading_empty = 0
    # Recompute: count leading empties in original header_vals from sheet
    original_header = [cell_to_str(ws.cell(header_row_idx, c).value) for c in range(1, ws.max_column + 1)]
    leading_empty = 0
    for v in original_header:
        if v == '':
            leading_empty += 1
        else:
            break
    header = [v for v in original_header[leading_empty:] if True]
    # Trim trailing empties of header
    while header and header[-1] == '':
        header.pop()
    # Collect data rows
    data = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        raw = [ws.cell(r, c).value for c in range(1 + leading_empty, 1 + leading_empty + len(header))]
        if all(v is None or str(v).strip() == '' for v in raw):
            continue
        data.append(raw)
    md = [f'# {sheet_name}\n']
    md.append(f'> Domain slug: `{slug}` · {len(data)} stories\n')
    md.append('## Stories\n')
    md.append(table_md(header, data))
    return '\n'.join(md)


def extract_business_flow(ws):
    """Master Business Flow index sheet."""
    rows = list(read_rows(ws))
    # Find header row with 'Nghiệp Vụ' or 'No'
    header_row_idx = None
    for r, row in rows:
        vals = [cell_to_str(v) for v in row]
        if 'No' in vals and ('Nhóm' in vals or 'Nghiệp Vụ' in vals):
            header_row_idx = r
            break
    md = ['# Business Flow — Master Index\n']
    md.append('> Tổng quan các nghiệp vụ của dự án ESKITCHEN (từ sheet `Business Flow`).\n')
    if header_row_idx is None:
        md.append('> Không nhận diện header — dump raw.')
        for r, row in rows:
            md.append(f'- R{r}: {" | ".join(cell_to_str(v) for v in row)}')
        return '\n'.join(md)
    original_header = [cell_to_str(ws.cell(header_row_idx, c).value) for c in range(1, ws.max_column + 1)]
    leading = 0
    for v in original_header:
        if v == '':
            leading += 1
        else:
            break
    header = original_header[leading:]
    while header and header[-1] == '':
        header.pop()
    data = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        raw = [ws.cell(r, c).value for c in range(1 + leading, 1 + leading + len(header))]
        if all(v is None or str(v).strip() == '' for v in raw):
            continue
        data.append(raw)
    md.append(f'## Danh sách {len(data)} nghiệp vụ\n')
    md.append(table_md(header, data))
    return '\n'.join(md)


def extract_function_list(ws):
    """Function List: metadata + SUMMARY + PHASE 1 + PHASE 2 detail rows."""
    md = ['# Function List\n']
    md.append('> Master list functions (từ sheet `Function List`). Nguồn: file ESTIMATION + REQUIREMENT.\n')
    rows = list(read_rows(ws))
    # 1. Header section (rows < 10)
    md.append('## Source Links\n')
    for r, row in rows:
        if r >= 10:
            break
        vals = [cell_to_str(v) for v in row if cell_to_str(v)]
        if vals:
            md.append('- ' + ' · '.join(vals))
    # 2. Summary table (rows 10..17 typically). Header at row 10, cols B..F.
    md.append('\n## Summary by Epic\n')
    summary = []
    sum_header = None
    for r, row in rows:
        if 10 <= r <= 18:
            # Detect leading empty cols then take next 5 (ID, EPIC, PHASE 1, PHASE 2, TOTAL)
            leading = 0
            for v in row:
                if v is None or str(v).strip() == '':
                    leading += 1
                else:
                    break
            block = [cell_to_str(v) for v in row[leading:leading + 5]]
            while block and block[-1] == '':
                block.pop()
            if not block:
                continue
            if block[0] == 'ID':
                sum_header = block
            else:
                summary.append(block)
    if sum_header and summary:
        md.append(table_md(sum_header, summary))
    # 3. Phase 1 table — header row contains 'STT' 'PHASE' 'ID' 'EPIC' 'STORY' 'DESCRIPTION'
    # Find header indices for phase blocks
    phase_headers = []
    for r, row in rows:
        vals = [cell_to_str(v) for v in row]
        if 'STT' in vals and 'EPIC' in vals and 'STORY' in vals:
            phase_headers.append((r, vals))
    for i, (hr, hvals) in enumerate(phase_headers):
        next_hr = phase_headers[i + 1][0] if i + 1 < len(phase_headers) else ws.max_row + 1
        # Determine leading empty cols
        leading = 0
        for v in hvals:
            if v == '':
                leading += 1
            else:
                break
        header = hvals[leading:]
        while header and header[-1] == '':
            header.pop()
        # Truncate to the actual phase columns (drop trailing summary overlap if any)
        # Header looks like [STT, PHASE, ID, EPIC, STORY, DESCRIPTION, CUSTOMER UAT, GOLIVE, '', ID, EPIC, PHASE 1, ...]
        # The summary overlap starts at the SECOND 'ID' — find it skipping the first occurrence.
        try:
            first_id = header.index('ID')
            second_id = header.index('ID', first_id + 1)
            header = header[:second_id]
            # Drop trailing empty column before the summary overlap (if any)
            while header and header[-1] == '':
                header.pop()
        except ValueError:
            pass
        data = []
        for r in range(hr + 1, next_hr):
            raw = [ws.cell(r, c).value for c in range(1 + leading, 1 + leading + len(header))]
            if all(v is None or str(v).strip() == '' for v in raw):
                continue
            data.append(raw)
        if not data:
            continue
        md.append(f'\n## Phase {i + 1} — {len(data)} rows\n')
        md.append(table_md(header, data))
    return '\n'.join(md)


def extract_overview(ws):
    """Overview sheet — empty/sparse."""
    rows = list(read_rows(ws))
    md = ['# Overview\n']
    if not rows:
        md.append('> Sheet `Overview` không có dữ liệu non-empty trong xlsx này.')
        return '\n'.join(md)
    for r, row in rows[:200]:
        vals = [cell_to_str(v) for v in row if cell_to_str(v)]
        if vals:
            md.append(f'- R{r}: ' + ' · '.join(vals))
    return '\n'.join(md)


def extract_screen_code_rule(ws):
    md = ['# Screen Code Rule\n']
    md.append('> Quy tắc đặt Screen Code cho dự án ESKITCHEN.\n')
    rows = list(read_rows(ws))
    for r, row in rows[:200]:
        # Indent by leading empty columns to preserve structure
        leading = 0
        for v in row:
            if v is None or str(v).strip() == '':
                leading += 1
            else:
                break
        content = [cell_to_str(v) for v in row[leading:] if cell_to_str(v)]
        if not content:
            continue
        prefix = '  ' * leading
        md.append(f'{prefix}- ' + ' · '.join(content))
    return '\n'.join(md)


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    OUT.mkdir(parents=True, exist_ok=True)
    (OUT / 'domains').mkdir(parents=True, exist_ok=True)
    counts = {}
    for name in wb.sheetnames:
        if name in SKIP:
            continue
        ws = wb[name]
        if name == 'Function List':
            content = extract_function_list(ws)
            path = OUT / 'function-list.md'
        elif name == 'Overview':
            content = extract_overview(ws)
            path = OUT / 'overview.md'
        elif name == 'Screen Code Rule':
            content = extract_screen_code_rule(ws)
            path = OUT / 'screen-code-rule.md'
        elif name == 'Business Flow':
            content = extract_business_flow(ws)
            path = OUT / 'business-flow-index.md'
        elif name in BF_SLUG:
            slug = BF_SLUG[name]
            content = extract_bf(ws, name, slug)
            path = OUT / 'domains' / f'{slug}.md'
        else:
            print(f'  ! Skipped unknown sheet: {name}', file=sys.stderr)
            continue
        path.write_text(content, encoding='utf-8')
        size = path.stat().st_size
        counts[str(path)] = size
        print(f'  ✓ {path}  ({size:,} bytes)')
    print(f'\nTotal: {len(counts)} files, {sum(counts.values()):,} bytes')


if __name__ == '__main__':
    main()
